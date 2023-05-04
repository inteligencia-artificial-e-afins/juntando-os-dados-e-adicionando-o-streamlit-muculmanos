from bs4 import BeautifulSoup    
from openpyxl import Workbook
import requests
import pandas as pd

url = "https://imobiliariaperez.com.br/alugar" # site para extrair
response = requests.get(url) #request html
soup = BeautifulSoup(response.content, 'html.parser') #configurao beautifulsoup
myInfoArray = []

links = soup.find_all('a',class_="slide-home-btn")

for link in links:
    pushHref = link['href']
    newResponse = requests.get(pushHref)
    newSoup = BeautifulSoup(newResponse.content, 'html.parser') #configurao beautifulsoup
    titleA = newSoup.find('h1', class_= "elementor-heading-title").get_text()
    value = newSoup.find('h2', class_= "elementor-heading-title").get_text() 
    description = newSoup.find_all('p')[3].get_text()
    arrayInfo =  newSoup.find_all('div', class_= "col-6")

    valueArea = arrayInfo[-1].text

    area = valueArea.split()[0] +""+ valueArea.split()[2]

    if value != 'Sob Consulta':
        value += ",00"

    title = titleA.split()[0]   

    aux = [title, value, area.strip(), description]
    myInfoArray.append(aux)

workbook = Workbook()

sheet = workbook.active

sheet['A1'] = 'Título'
sheet['B1'] = 'Aluguel'
sheet['C1'] = 'Área'
sheet['D1'] = 'Descrição'

for i, row in enumerate(myInfoArray):
    sheet.cell(row=i+2, column=1, value=row[0])
    sheet.cell(row=i+2, column=2, value=row[1])
    sheet.cell(row=i+2, column=3, value=row[2])
    sheet.cell(row=i+2, column=4, value=row[3])
    
workbook.save(filename='scrapImobiliaria.xlsx')

df = pd.read_excel('scrapImobiliaria.xlsx')

df.to_csv('scrapImobiliaria.csv', index=False)
