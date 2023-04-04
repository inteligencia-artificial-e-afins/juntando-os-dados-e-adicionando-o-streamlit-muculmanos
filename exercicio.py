from bs4 import BeautifulSoup
from openpyxl import Workbook
import requests


url = "https://imobiliariaperez.com.br/alugar" # site para extrair
response = requests.get(url) #request html


soup = BeautifulSoup(response.content, 'html.parser') #configurao beautifulsoup

myInfoArray = []

links = soup.find_all('a',class_="slide-home-btn")

for link in links:
    pushHref = link['href']
    newResponse = requests.get(pushHref)
    newSoup = BeautifulSoup(newResponse.content, 'html.parser') #configurao beautifulsoup
    title = newSoup.find('h1', class_= "elementor-heading-title").get_text()
    value = newSoup.find('h2', class_= "elementor-heading-title").get_text()
    area =  newSoup.find('div', class_= "col-6").get_text()
    description = newSoup.find_all('p')[3].get_text()
    address = "londrina-PR"
    aux = [title, value, area.strip(), description, address ]
    myInfoArray.append(aux)

workbook = Workbook()

sheet = workbook.active

sheet['A1'] = 'Título'
sheet['B1'] = 'Endereço'
sheet['C1'] = 'Aluguel'
sheet['D1'] = 'Área'
sheet['E1'] = 'Descrição'

# Adiciona os dados
for i, row in enumerate(myInfoArray):
    sheet.cell(row=i+2, column=1, value=row[0])
    sheet.cell(row=i+2, column=2, value=row[4])
    sheet.cell(row=i+2, column=3, value=row[1])
    sheet.cell(row=i+2, column=4, value=row[2])
    sheet.cell(row=i+2, column=5, value=row[3])

# Salva o arquivo
workbook.save(filename='scrapImobiliarias2.xlsx')